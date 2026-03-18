import os
import PyPDF2
import openpyxl


def parse_uploaded_file(filepath: str) -> dict:
    """
    Parse a user-uploaded file and return its content in a structure the agent can reason over.

    Dispatches to the appropriate parser based on file extension.
    Supports: .pdf (text extraction), .xlsx/.xls (table extraction), .txt/.csv (raw read).

    The agent receives this dict and reasons over its 'content' field.
    For PDFs, it reasons over the extracted text directly.
    For Excel files, it reasons over the 2D table structure.
    The 'char_count' field helps the agent decide if the content is too long to include
    in full and whether to summarize or chunk.

    Args:
        filepath: Absolute path to the file in the uploads/ directory.

    Returns:
        Dict with keys: type, filename, content, char_count.
        On any parsing error, returns {"type": "error", "filename": ..., "content": str(error)}.
    """
    # TODO: ext = os.path.splitext(filepath)[1].lower()
    # TODO: filename = os.path.basename(filepath)
    # TODO: Wrap in try/except. On exception: return {"type": "error", "filename": filename, "content": str(e)}
    # TODO: Dispatch:
    #         if ext == ".pdf":        content = _extract_pdf_text(filepath),   type_ = "pdf"
    #         elif ext in (".xlsx", ".xls"): content = _extract_xlsx_tables(filepath), type_ = "xlsx"
    #         else:                    content = open(filepath, "r", errors="ignore").read(), type_ = "text"
    # TODO: Compute char_count:
    #         If content is a string: char_count = len(content)
    #         If content is a dict (xlsx): char_count = sum of lengths of all string cell values.
    # TODO: Return {"type": type_, "filename": filename, "content": content, "char_count": char_count}
    pass


def _extract_pdf_text(filepath: str) -> str:
    """
    Extract all text from a PDF file as a single concatenated string.

    Pages are joined with newlines. Pages with no extractable text (scanned images
    without OCR) contribute an empty string — no error is raised.
    For scanned PDFs, consider adding pytesseract OCR support in Phase 3.

    Args:
        filepath: Path to the PDF file.

    Returns:
        Full extracted text. May be empty for image-only PDFs.
    """
    # TODO: Open filepath in "rb" mode.
    # TODO: reader = PyPDF2.PdfReader(file_handle)
    # TODO: pages = [page.extract_text() or "" for page in reader.pages]
    # TODO: Return "\n".join(pages)
    pass


def _extract_xlsx_tables(filepath: str) -> dict[str, list[list]]:
    """
    Extract all sheets from an Excel file as a dict of 2D lists.

    Each sheet becomes a key. The value is a list of rows where each row
    is a list of cell values as plain Python types.
    Empty rows (all None values) are skipped to reduce noise.

    This structure lets the agent reason over financial tables the user has prepared —
    for example, a personal portfolio tracker or a custom financial model.

    Args:
        filepath: Path to the .xlsx file.

    Returns:
        {"Sheet1": [[val, val, ...], [val, val, ...]], "Sheet2": [...]}
        Cell values are str, int, float, or None. No pandas or numpy types.
    """
    # TODO: wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    # TODO: result = {}
    # TODO: For each sheet_name in wb.sheetnames:
    #         ws = wb[sheet_name]
    #         rows = []
    #         For each row in ws.iter_rows():
    #           row_values = [cell.value for cell in row]
    #           If any(v is not None for v in row_values):  ← skip fully empty rows
    #             rows.append(row_values)
    #         result[sheet_name] = rows
    # TODO: Return result
    pass
